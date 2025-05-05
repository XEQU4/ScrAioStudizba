import logging
from typing import List

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.fills import PatternFill
from openpyxl.workbook import Workbook


async def create_excel_file(name: str) -> None:
    """
    Creates an Excel file with predefined column headers and styles.
    If the file already exists, it will be opened without overwriting.

    :param name: File name or path to the Excel file
    """
    columns = [
        'Название Кафедры',
        'ФИО преподавателя',
        'Описание преподавателя (HTML)',
        'Описание преподавателя (TEXT)',
        'Ссылки преподавателя',
        'Фото преподавателя'
    ]
    widths = [30, 50, 100, 100, 70, 70]

    try:
        wb = load_workbook(name)  # Try to open existing file
    except Exception:
        wb = Workbook()  # Create a new workbook if it doesn't exist

    ws = wb.active

    # Set column headers with styling
    for i, (col_name, width) in enumerate(zip(columns, widths), start=1):
        cell = ws.cell(row=1, column=i, value=col_name)
        cell.font = Font(size=10, color='FFFFFF', bold=True)
        cell.fill = PatternFill(fgColor="4F81BD", fill_type="solid")
        col_letter = chr(64 + i)  # Convert index to Excel column letter (A, B, C, ...)
        ws.column_dimensions[col_letter].width = width

    # Save the file
    try:
        wb.save(name)
    except PermissionError:
        logging.error("Excel file is open. Please close it and try again.")
        return


async def write_datas(data: List[List], path: str) -> None:
    """
    Appends rows of data to an existing Excel file.

    :param data: List of rows, where each row is a list of values
    :param path: Path to the Excel file
    """
    try:
        wb = load_workbook(path)
        ws = wb.active

        for row in data:
            ws.append(row)

        wb.save(path)

    except Exception as e:
        logging.error(f"Failed to write data to {path}: {e}")
